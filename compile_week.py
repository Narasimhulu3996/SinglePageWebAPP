#!/usr/bin/env python3
import sys
import os
import json
import openpyxl

canonical_stations = [
    'Punganur UPS', 'Madanapalle Rural UPS', 'Sodam', 'Madanapalle II Town UPS',
    'B.Kothakota UPS', 'Ramasamudram', 'Nimmanapalle', 'P.T.Samudram (PTM)',
    'Mulakalacheruvu', 'Mudiveedu', 'Thamballapalli', 'Somala', 'Peddamandyam',
    'Madanapalle I Town UPS', 'Chowdepalli', 'Kalikiri UPS', 'Rayachoty UPS',
    'Lakkireddipalli', 'Kalakada', 'Piler UPS', 'Voyalpad', 'Ramapuram',
    'Rayachoty Traffic', 'K.V.Palli', 'Galiveedu', 'Gurramkonda', 'Sambepalli',
    'Chinnamandem'
]

def norm(name):
    if name is None: return ''
    n = ' '.join(str(name).split()).lower()
    n = n.replace(' ps', '').replace(' ups', '').replace(',', '').replace('.', '').replace('-', '').strip()
    if 'rayachoty traffic' in n:
        return 'rayachoty traffic'
    elif 'rayachoty' in n:
        return 'rayachoty'
    return n

norm_canon = {norm(s): s for s in canonical_stations}

def get_row_by_station(ws, name_col, target_norm):
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[2:]:
        if len(r) > name_col:
            val = r[name_col]
            if val is None:
                continue
            if norm(val) == target_norm:
                return r
    return None

def get_sheet_and_col(wb, sheet_name_cols, key):
    # Find worksheet whose name contains key case-insensitively
    for name in wb.sheetnames:
        if key.lower() in name.lower():
            return wb[name], sheet_name_cols[name]
    # Fallback to exact match to raise original KeyError if not found
    return wb[key], sheet_name_cols[key]

def find_column_by_name(ws, keywords, default_idx):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return default_idx
    
    # Try looking in row 2 (header)
    for c_idx, val in enumerate(rows[1]):
        if val is not None:
            val_clean = ' '.join(str(val).split()).lower()
            if any(k in val_clean for k in keywords):
                return c_idx
                
    # Fallback to row 1
    for c_idx, val in enumerate(rows[0]):
        if val is not None:
            val_clean = ' '.join(str(val).split()).lower()
            if any(k in val_clean for k in keywords):
                return c_idx
                
    return default_idx

def main():
    wb = openpyxl.load_workbook('PS wise DATA 04062026-10062026.xlsx', data_only=True)
    
    # Load previous week's JSON data to carry over CEIR Score and eOffice if needed
    prev_data = {}
    prev_json_path = os.path.join('data', 'weeks', 'ANM_PS_03_06_2026-09_06_2026.json')
    if os.path.exists(prev_json_path):
        try:
            with open(prev_json_path, 'r', encoding='utf-8') as f:
                prev_json = json.load(f)
                for st in prev_json.get('stations', []):
                    prev_data[norm(st['name'])] = {
                        'CEIR Score': st['parameters'].get('CEIR Score', 50.0),
                        'eOffice': st['parameters'].get('eOffice', 50.0)
                    }
            print('Successfully loaded previous week data for carry over.')
        except Exception as e:
            print('Warning: could not load previous week JSON:', e)
            
    # Find name column index in each sheet
    sheet_name_cols = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            sheet_name_cols[sheet] = -1
            continue
        name_col = -1
        for c_idx, val in enumerate(rows[1]):
            if val is not None and any(k in str(val).lower() for k in ['ps', 'station', 'unit name']):
                name_col = c_idx
                break
        sheet_name_cols[sheet] = name_col

    # Columns of consolidated sheet
    columns = [
        'District', 'MedLEaPR PME', 'MedLEaPR MLC', 'NATGRID ', 'DGP Dashboard', 
        'CCTNS Overall', 'CCTNS IIF 1-7', 'CCTNS IIF 8-15', 'Zero FIRs', 'Citizen Portal', 
        'eSakshya SID%', 'CriMAC ', 'CEIR Score', 'eProsecution', 'FIRs & CS Con', 
        'eOffice', 'APOLIS ', 'Drone Flying', 'Dial112 Tabs', 'Dial112 Resp', 
        'CCTV360 Usage', 'News360AI'
    ]
    
    weights = ['Wt →', 0.05, 0.05, 0.05, 0.05, 0.05, 0.025, 0.025, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05]
    
    compiled_rows = []
    
    for station in canonical_stations:
        st_norm = norm(station)
        row_data = [station] # First column is station name
        
        # 1. MedLEaPR PME
        sheet, col = get_sheet_and_col(wb, sheet_name_cols, 'MedLEaPR-PMR')
        raw_col = find_column_by_name(sheet, ['raw value', 'rawvalue'], 4)
        r = get_row_by_station(sheet, col, st_norm)
        pme = float(r[raw_col]) if (r and len(r) > raw_col and r[raw_col] is not None) else 100.0
        row_data.append(pme)
        
        # 2. MedLEaPR MLC
        sheet, col = get_sheet_and_col(wb, sheet_name_cols, 'MedLEaPR-MLC')
        raw_col = find_column_by_name(sheet, ['raw value', 'rawvalue'], 4)
        r = get_row_by_station(sheet, col, st_norm)
        mlc = float(r[raw_col]) if (r and len(r) > raw_col and r[raw_col] is not None) else 100.0
        row_data.append(mlc)
        
        # 3. NATGRID
        row_data.append(0.0)
        
        # 4. DGP Dashboard
        sheet, col = get_sheet_and_col(wb, sheet_name_cols, 'DGP Dashboard')
        raw_col = find_column_by_name(sheet, ['raw value', 'rawvalue'], 3)
        r = get_row_by_station(sheet, col, st_norm)
        dgp = float(r[raw_col]) if (r and len(r) > raw_col and r[raw_col] is not None) else 100.0
        row_data.append(dgp)
        
        # 5. CCTNS Overall
        sheet, col = get_sheet_and_col(wb, sheet_name_cols, 'IIFS-TILLDATE')
        raw_col = find_column_by_name(sheet, ['raw value', 'rawvalue'], 15)
        r = get_row_by_station(sheet, col, st_norm)
        cctns_ovr = float(r[raw_col]) if (r and len(r) > raw_col and r[raw_col] is not None) else 0.0
        row_data.append(cctns_ovr)
        
        # 6. CCTNS IIF 1-7
        sheet, col = get_sheet_and_col(wb, sheet_name_cols, 'IIF1-7')
        raw_col = find_column_by_name(sheet, ['raw value', 'rawvalue'], 13)
        r = get_row_by_station(sheet, col, st_norm)
        iif1_7 = float(r[raw_col]) if (r and len(r) > raw_col and r[raw_col] is not None) else 0.0
        row_data.append(iif1_7)
        
        # 7. CCTNS IIF 8-15
        sheet, col = get_sheet_and_col(wb, sheet_name_cols, 'IIF8-15')
        raw_col = find_column_by_name(sheet, ['raw value', 'rawvalue'], 12)
        r = get_row_by_station(sheet, col, st_norm)
        iif8_15 = float(r[raw_col]) if (r and len(r) > raw_col and r[raw_col] is not None) else 0.0
        row_data.append(iif8_15)
        
        # 8. Zero FIRs
        sheet, col = get_sheet_and_col(wb, sheet_name_cols, 'ZERO FIR')
        raw_col = find_column_by_name(sheet, ['raw value', 'rawvalue'], 6)
        r = get_row_by_station(sheet, col, st_norm)
        zero_fir = float(r[raw_col]) if (r and len(r) > raw_col and r[raw_col] is not None) else 100.0
        row_data.append(zero_fir)
        
        # 9. Citizen Portal
        sheet, col = get_sheet_and_col(wb, sheet_name_cols, 'Petition')
        raw_col = find_column_by_name(sheet, ['raw value', 'rawvalue'], 5)
        r = get_row_by_station(sheet, col, st_norm)
        cp = float(r[raw_col]) if (r and len(r) > raw_col and r[raw_col] is not None) else 100.0
        row_data.append(cp)
        
        # 10. eSakshya SID%
        sheet, col = get_sheet_and_col(wb, sheet_name_cols, 'eSakshya')
        raw_col = find_column_by_name(sheet, ['raw value', 'rawvalue'], 8)
        r = get_row_by_station(sheet, col, st_norm)
        esak = float(r[raw_col]) if (r and len(r) > raw_col and r[raw_col] is not None) else 100.0
        row_data.append(min(100.0, esak))
        
        # 11. CriMAC
        sheet, col = get_sheet_and_col(wb, sheet_name_cols, 'Cri.MAC')
        raw_col = find_column_by_name(sheet, ['raw value', 'rawvalue'], 11)
        r = get_row_by_station(sheet, col, st_norm)
        crimac = float(r[raw_col]) if (r and len(r) > raw_col and r[raw_col] is not None) else 100.0
        row_data.append(crimac)
        
        # 12. CEIR Score
        ceir = prev_data.get(st_norm, {}).get('CEIR Score', 50.0)
        row_data.append(ceir)
        
        # 13. eProsecution
        sheet, col = get_sheet_and_col(wb, sheet_name_cols, 'eProsecution')
        raw_col = find_column_by_name(sheet, ['raw value', 'rawvalue'], 5)
        r = get_row_by_station(sheet, col, st_norm)
        epros = float(r[raw_col]) if (r and len(r) > raw_col and r[raw_col] is not None) else 0.0
        row_data.append(epros)
        
        # 14. FIRs & CS Con
        sheet, col = get_sheet_and_col(wb, sheet_name_cols, 'FIRs&CS Consm in Court')
        raw_col = find_column_by_name(sheet, ['raw value', 'rawvalue'], 10)
        r = get_row_by_station(sheet, col, st_norm)
        f_j = float(r[raw_col]) if (r and len(r) > raw_col and r[raw_col] is not None) else 0.0
        row_data.append(f_j)
        
        # 15. eOffice
        eoff = prev_data.get(st_norm, {}).get('eOffice', 50.0)
        row_data.append(eoff)
        
        # 16. APOLIS
        sheet, col = get_sheet_and_col(wb, sheet_name_cols, 'APOLIS')
        raw_col = find_column_by_name(sheet, ['raw value', 'rawvalue'], 9)
        r = get_row_by_station(sheet, col, st_norm)
        ap = float(r[raw_col]) if (r and len(r) > raw_col and r[raw_col] is not None) else 100.0
        row_data.append(ap)
        
        # 17. Drone Flying
        sheet, col = get_sheet_and_col(wb, sheet_name_cols, 'DroneFlying')
        raw_col = find_column_by_name(sheet, ['raw value', 'rawvalue'], 6)
        r = get_row_by_station(sheet, col, st_norm)
        df = float(r[raw_col]) if (r and len(r) > raw_col and r[raw_col] is not None) else 100.0
        row_data.append(df)
        
        # 18. Dial112 Tabs
        sheet, col = get_sheet_and_col(wb, sheet_name_cols, 'Dail-112 TABs')
        raw_col = find_column_by_name(sheet, ['raw value', 'rawvalue'], 5)
        r = get_row_by_station(sheet, col, st_norm)
        d_tabs = float(r[raw_col]) if (r and len(r) > raw_col and r[raw_col] is not None) else 100.0
        row_data.append(d_tabs)
        
        # 19. Dial112 Resp
        sheet, col = get_sheet_and_col(wb, sheet_name_cols, 'Dail-112 Response')
        raw_col = find_column_by_name(sheet, ['raw value', 'rawvalue'], 6)
        r = get_row_by_station(sheet, col, st_norm)
        d_resp = float(r[raw_col]) if (r and len(r) > raw_col and r[raw_col] is not None) else 100.0
        row_data.append(d_resp)
        
        # 20. CCTV360 Usage
        sheet, col = get_sheet_and_col(wb, sheet_name_cols, 'CCTV360')
        raw_col = find_column_by_name(sheet, ['raw value', 'rawvalue'], 18)
        r = get_row_by_station(sheet, col, st_norm)
        cctv = float(r[raw_col]) if (r and len(r) > raw_col and r[raw_col] is not None) else 0.0
        row_data.append(cctv)
        
        # 21. News360AI
        row_data.append(0.0)
        
        compiled_rows.append(row_data)
        
    # Create output workbook
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = 'Sheet1'
    
    # Write header
    out_ws.append(columns)
    
    # Write weights
    out_ws.append(weights)
    
    # Write station rows
    for row in compiled_rows:
        out_ws.append(row)
        
    out_file = os.path.join('data', 'weeks', 'ANM_PS_04_06_2026-10_06_2026.xlsx')
    out_wb.save(out_file)
    print(f'Successfully compiled and saved workbook to {out_file}')

if __name__ == '__main__':
    main()
